import openpyxl, os, csv

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
    	continue
    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        csvName = excelFile[:-4] + '_' + sheetName
        csvFile  = open(csvName, 'w', newline='')
        csvWriter = csv.writer(csvFile)

        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            for colNum in range(1, sheet.max_column + 1):
            	data = sheet.cell(row=rowNum,column=colNum).value
            	rowData.append(data)
            for row in rowData:
            	csvWriter.writerow(rowData)

        csvFile.close()