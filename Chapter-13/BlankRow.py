import openpyxl,sys

N = int(sys.argv[1])
M = int(sys.argv[2])
filename = sys.argv[3]
wb = openpyxl.load_workbook(filename)
sheet = wb['Sheet1']

max_col = sheet.max_column
max_row = sheet.max_row

data= []
for col in range(1,max_col+1):
	for row in range(1, max_row+1):
		Before = sheet.cell(row=row, column = col).value
		data.append(Before)

print(data)
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

i=0
for col in range(1,max_col+1):
	for row in range(1, max_row+M+1):
		if row >=N and row<N+M:
			values = ''
			sheet2.cell(row=row, column=col).value = values
		else:
			sheet2.cell(row = row, column=col).value = data[i]
			i+=1

wb2.save('blank.xlsx')
