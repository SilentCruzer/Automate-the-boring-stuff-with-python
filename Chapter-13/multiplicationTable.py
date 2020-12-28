import openpyxl, sys
from openpyxl.styles import Font

number = sys.argv[1]
wb = openpyxl.Workbook()
sheet = wb['Sheet']
bold=Font(size = 14 , bold=True)

for col in range(1,int(number)+1):
	for row in range(1,int(number)+1):
		if row == 1:
			cell = sheet.cell(row=1, column=col+1)
			sheet.cell(row=1, column=col+1).font = bold
			cell.value = col
		if col == 1:
			cell = sheet.cell(row=row+1, column=1)
			sheet.cell(row=row+1, column=1).font = bold
			cell.value = row
		cell = sheet.cell(row=row+1, column=col+1)
		cell.value = (row) * (col)

wb.save('MultiplicationTable.xlsx')

