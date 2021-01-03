import openpyxl,sys

excel_file = sys.argv[1]
wb = openpyxl.load_workbook(excel_file)
sheet = wb['Sheet1']
col_size = sheet.max_column
row_size = sheet.max_row	
row,col = 1,1

while col <= col_size:
	filename = 'text' + str(col)
	file = open(filename,'w')
	while row<= row_size:
		string = sheet.cell(row=row,column=col).value
		file.write(str(string) + '\n')
		row+=1
	file.close()
	col+=1
	row=1






