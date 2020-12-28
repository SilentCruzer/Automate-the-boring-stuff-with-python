import openpyxl

wb = openpyxl.load_workbook('BeforeInverted.xlsx')
sheet = wb['Sheet1']

nb = openpyxl.Workbook()
sheet2 = nb['Sheet']
col_size  = sheet.max_column
row_size = sheet.max_row

for col in range(col_size,0,-1):
	for row in range(row_size,0,-1):
		cell = sheet2.cell(row=col, column=row)
		invert = sheet.cell(row=row, column=col)
		cell.value = invert.value

nb.save('AfterInverted.xlsx')