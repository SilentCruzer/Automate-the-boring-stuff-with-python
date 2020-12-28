import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']
print(sheet.cell(row=1,column=1).value)